import math
import openpyxl
# from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.styles.colors import Color
from decimal import Decimal


IN_XLSX_FILE = "Кабельный журнал.xlsx"
OUT_XLSX_FILE = "Таблица прокладки кабеля.xlsx"

CSS_NAMES = {
    "-": "По конструкциям",
    "Не разложен": "По конструкциям",
    "S5": "Лоток",
    "U5": "Лоток",
    "L5": "Лоток",
    "Л ": "Гофра",
    "Т ": "Гофра",
    "СТ ": "Гофра",
    "ЛО ": "Гофра",
    "ЛЧ ": "Гофра",
    "ТО ": "Гофра",
    "ЛЖ ": "Труба",
    "ТЖ ": "Труба",
    "ДГ ": "Труба",
    "ДЖ ": "Труба",
    "100х50": "Короб",
}


class Cable:
    def __init__(self, cable_grade, conductor_cross_section):
        self.__set_cable_grade(cable_grade)
        self.__set_conductor_count(conductor_cross_section)
        self.__set_conductor_cross_section(conductor_cross_section)

    def get_cable_grade(self):
        return self.__cable_grade

    def __set_cable_grade(self, cable_grade):
        self.__cable_grade = cable_grade.replace('-0.66', '')

    cable_grade = property(get_cable_grade, __set_cable_grade)

    def get_conductor_count(self):
        return self.__conductor_count

    def __set_conductor_count(self, conductor_cross_section):
        self.__conductor_count = int(conductor_cross_section.split("x")[0])

    conductor_count = property(get_conductor_count, __set_conductor_count)

    def get_conductor_cross_section(self):
        return self.__conductor_cross_section

    def __set_conductor_cross_section(self, conductor_cross_section):
        self.__conductor_cross_section = Decimal(conductor_cross_section.split("x")[1].replace(",", "."))

    conductor_cross_section = property(get_conductor_cross_section, __set_conductor_cross_section)

    def __str__(self):
        return f"{self.cable_grade} {self.conductor_count}х{str(self.conductor_cross_section).replace('.', ',')} мм²"

    def __lt__(self, other):
        if self.cable_grade < other.cable_grade:
            return True
        if (self.cable_grade == other.cable_grade and
                self.conductor_count < other.conductor_count):
            return True
        if (self.cable_grade == other.cable_grade and
                self.conductor_count == other.conductor_count and
                self.conductor_cross_section < other.conductor_cross_section):
            return True
        return False

    def __gt__(self, other):
        if self.cable_grade > other.cable_grade:
            return True
        if (self.cable_grade == other.cable_grade and
                self.conductor_count > other.conductor_count):
            return True
        if (self.cable_grade == other.cable_grade and
                self.conductor_count == other.conductor_count and
                self.conductor_cross_section > other.conductor_cross_section):
            return True
        return False

    def __eq__(self, other):
        return (self.cable_grade == other.cable_grade and
                self.conductor_count == other.conductor_count and
                self.conductor_cross_section == other.conductor_cross_section)

    def __hash__(self):
        return hash(str(self))


class CSS:
    def __init__(self, name, dimension=None):
        self.name = name
        self.__set_css_type(name)
        self.dimension = dimension

    def get_css_type(self):
        return self.css_type

    def __set_css_type(self, name):
        for css_name in CSS_NAMES:
            if name.startswith(css_name):
                self.css_type = CSS_NAMES[css_name]
                # if CSS_NAMES[css_name] == "По конструкциям":
                #     self.css_type = CSS_NAMES[css_name]
                # else:
                #     self.css_type = f"{CSS_NAMES[css_name]} {self.css_type}"

    css_type = property(get_css_type, __set_css_type)

    def __str__(self):
        # css = css_name[i]
        # if css_name[i] == "Лоток" or css_name[i] == "Короб":
        #     css = f"{css_name[i]} {css_size[i]} мм"
        # elif css_name[i] == "Гофра" or css_name[i] == "Труба":
        #     css = f"{css_name[i]} d.{css_size[i]} мм"
        # else:
        #     css = css_name[i]

        if self.css_type == "По конструкциям":
            return f"{self.css_type} {self.name}"
        else:
            return self.css_type


class Line:
    def __init__(self, cable, css, length):
        self.cable = cable
        self.css = css
        self.length = length
        self.cable_length_surplus = None


# def get_key(d, value):
#     for k, v in d.items():
#         if v == value:
#             return k


# def write_csv(out_csv_file, table, delimiter=";"):
#     # csv.register_dialect(";", delimiter=";", quoting=csv.QUOTE_NONE)
#     csv.register_dialect(delimiter, delimiter=delimiter, quoting=csv.QUOTE_NONE)
#
#     try:
#         with open(out_csv_file, "w", encoding="utf-8", newline="") as f:
#             f_writer = csv.writer(f, delimiter)
#             f_writer.writerows(table)
#     except PermissionError:
#         print(f"Открыт файл: '{out_csv_file}'. Заккройте файл и повторите операцию.")


def write_xlsx(out_xlsx_file, table):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Таблица прокладки кабеля"

    for row in table:
        ws.append(list(map(str, row)))

    rows = ws.max_row
    cols = ws.max_column

    # tab = openpyxl.worksheet.table.Table(displayName="Table1", ref=f"A1:{COLUMN_NAMES[cols - 1]}{rows}")
    # style = TableStyleInfo(name="TableStyleMedium9",
    #                        showFirstColumn=False,
    #                        showLastColumn=False,
    #                        showRowStripes=True,
    #                        showColumnStripes=True)
    # tab.tableStyleInfo = style
    # ws.add_table(tab)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 10
    ws.row_dimensions[1].height = 50

    for column in range(1, cols + 1):
        cell = ws.cell(row=1, column=column)
        cell.font = Font(name="Calibri", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        cell.fill = PatternFill(patternType="solid", fgColor=Color(rgb="6495ed"))

    for row in range(1, rows + 1):
        for column in range(1, cols + 1):
            cell = ws.cell(row=row, column=column)
            bd = Side(style="thin", color="000000")
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            if row != 1 and column != 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for column in range(1, cols + 1):
        cell = ws.cell(row=rows, column=column)
        cell.font = Font(name="Calibri", bold=True)

    try:
        wb.save(out_xlsx_file)
    except PermissionError:
        print(f"Открыт файл: '{out_xlsx_file}'. Заккройте файл и повторите операцию.")


def main():
    data_list = [["css_name",
                  "css_length",
                  "cable_name",
                  "cable_length",
                  "cable_length_surplus",
                  ]]
    css_name_set = set()  # Множество типов КНС (Гофра d.25 мм)
    cable_name_set = set()  # Множество типов кабеля (ВВГнг(А)-LS 5х1,5)

    wb = openpyxl.load_workbook(IN_XLSX_FILE)
    ws = wb.active  # Получаем активный лист
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] == "Позиция":
            continue
        css_name = row[3].replace("_x000D_", "").replace("\n", "").split(";")
        if not css_name[0]:
            continue
        # css_size = row[4].replace("_x000D_", "").replace("\n", "").split(";")
        css_length = [Decimal(i).quantize(Decimal("1.0000"))
                      for i in row[5].replace("_x000D_", "").replace("\n", "").replace(",", ".").split(";")]
        cable_grade = row[6]
        cable_cross_section = row[7]
        cable_length = Decimal(row[8].replace("_x000D_", "").replace(",", "."))
        cable_length_surplus = Decimal(cable_length)

        one_line = []  # Массив для участков одной линии
        css_one_line_set = set()  # Множество КНС для линии
        for i in range(len(css_name)):  # Парсим строку (разбираем по способам прокладки)
            cable_length_surplus -= css_length[i]

            for name in CSS_NAMES:
                if css_name[i].startswith(name):
                    css_name[i] = CSS_NAMES[name]
                    # if CSS_NAMES[name] == "По конструкциям":
                    #     css_name[i] = CSS_NAMES[name]
                    # else:
                    #     css_name[i] = f"{CSS_NAMES[name]} {css_name[i]}"

            css = css_name[i]
            # if css_name[i] == "Лоток" or css_name[i] == "Короб":
            #     css = f"{css_name[i]} {css_size[i]} мм"
            # elif css_name[i] == "Гофра" or css_name[i] == "Труба":
            #     css = f"{css_name[i]} d.{css_size[i]} мм"
            # else:
            #     css = css_name[i]

            cable = Cable(cable_grade, cable_cross_section)

            if i != 0:
                cable_length = 0

            if css in css_one_line_set:
                for section in one_line:
                    if section[0] == css:
                        section[1] += css_length[i]
            else:
                one_line.append([css,
                                 css_length[i],
                                 cable,
                                 cable_length,
                                 cable_length_surplus])

            css_one_line_set.add(css)
            css_name_set.add(css)
            cable_name_set.add(cable)

        if cable_length_surplus != 0:
            for section in one_line:
                if section[0] == "По конструкциям":
                    section[1] += cable_length_surplus

        data_list.extend(one_line)

    cable_dict = {}  # Словарь с длинами кабелей по маркам

    for section in data_list[1:]:
        if section[3] != 0:
            if section[2] not in cable_dict.keys():
                cable_dict[section[2]] = section[3]
            else:
                cable_dict[section[2]] += section[3]

    summary_table = []

    fields = ["Марка кабеля", "Количество кабеля по спецификации, м"]
    fields.extend(sorted(list(css_name_set)))
    fields.extend(["Сумма", "Проверка"])

    summary_table.append(fields)

    for cable in sorted(list(cable_name_set)):

        for line in data_list[1:]:

            if not line[2] == cable:
                continue

            for row in summary_table[1:]:
                if row[0] == line[2]:
                    row[summary_table[0].index(line[0])] += line[1]
                    break
            else:
                summary_line = [line[2], cable_dict[line[2]], ]
                for field in fields[2:-2]:
                    if line[0] == field:
                        summary_line.append(line[1])
                    else:
                        summary_line.append(0)
                summary_table.append(summary_line)

    for i in range(len(summary_table[1:])):
        total_css_length = sum(summary_table[i + 1][2:]).quantize(Decimal("1.000"))
        summary_table[i + 1].append(total_css_length)
        summary_table[i + 1].append(summary_table[i + 1][1] == total_css_length)

    tmp_table = []
    for row in summary_table[1:]:
        tmp_table.append(row[2:-1])

    summary = ["Итого", "", ]
    summary.extend([str(n).replace(".", ",") for n in map(sum, zip(*tmp_table))])

    for i in range(len(summary_table[1:])):
        for j in range(len(summary_table[i+1][1:-1])):
            if j:
                cell = summary_table[i + 1][j + 1]
            else:
                cell = math.ceil(summary_table[i + 1][j + 1])
            cell = str(cell).replace(".", ",")
            if cell == "0":
                cell = ""
            summary_table[i+1][j+1] = cell
    summary_table.append(summary)

    write_xlsx(OUT_XLSX_FILE, summary_table)


if __name__ == "__main__":
    main()
